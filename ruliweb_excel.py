# 최종으로 모두 완성한 진짜 크롤링 코드 입니다.


import os
from ssl import Options
import sys
import re
import csv
import time
import traceback
from time import sleep
from datetime import datetime
import requests
from selenium import webdriver
import chromedriver_autoinstaller
from bs4 import BeautifulSoup as bs
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from selenium.webdriver.chrome.options import Options

# 새로운 모듈 추가
from operator import le
from openpyxl import Workbook
from openpyxl import load_workbook
import threading

# 게시글 id 들고오기 
# article_num, newurl, article_title, product_title, link, price
def hotdeal():
    global driver
    global worksheet_reset
    global load_wb
    
    global page
    page = 1
    
    print('루리웹 서칭 시작')
    
    final_matching_list = []   
    product_title_list = []
    price_list = []
    link_list = []
    
    matching_number_list = []
    article_title_list = []
    newurl_list = []
    stop = 0
    
    global searh_num, cell_num
    # searh_num = 61532
    # searh_num = int(searh_num)
    
    print('루리웹 서칭 시작')
    
    a=0
    for page_ in range(1,5):  
        # 기본경로 설정  
        base_url = "https://bbs.ruliweb.com/market/board/1020?page="+str(page_)
        html = requests.get(base_url)
        soup = bs(html.text, "html.parser")
        # driver.implicitly_wait(5)
        
        for i in range(0,28):
            # id
            count_a = len(soup.find_all(class_="table_body notice"))
            count_b = len(soup.find_all(class_="table_body notice inside"))
            count_c = len(soup.find_all(class_="table_body best inside"))
            
            count_all = count_a+count_b+count_c
            
            elements = soup.select('tr.table_body > td.id')[count_all+i]
            article_num = elements.text
            article_num = article_num.replace(" ","") 
            article_num = article_num.replace("\n","")
            
            print("매칭넘버는"+article_num+" 99")
            article_num = int(article_num)
            matching_number = article_num
            
            if page_ == 1:
                    # print('3')
                    if i == 0:
                        print("매칭넘버는"+str(matching_number)+"요거이다")
                        reset_num = int(matching_number)
                        last_num = int(reset_num) +1
                        print('현재의 최신 넘버',reset_num)
            if matching_number == int(searh_num):
                # print('4')
                stop = 1
                break
                '''하한:직전 크롤링에서의 마지막 작업 ID, 상한:이번 크롤링에서의 최신 작업 ID'''
            if matching_number > int(searh_num):
                '''다음페이지에서의 리셋문제 해결'''
                
                if (matching_number>=last_num):
                    print(matching_number,last_num,'매칭넘버가 더 크거나 같다, 이것만 스킵')
                    continue
                else:
                    last_num = matching_number
                
                
                    final_matching_list.append(matching_number) 
            else:
                stop = 1
                print('하한보다 작기때문에 스탑')
                break
            
            
            # 구매처, 제목
            # * table_body 자체이름을 비교하는 코드 추가하기 
            article_title = soup.select('tr.table_body > td.subject > div.relative > a.deco')[i].text
            article_title = article_title.replace(" ","") #space제거
            
            product_title = article_title[article_title.find(']')+1:]
            # article_title = re.sub("쿠팡|위메프|옥x|핫딜|롯데온|무배|지마켓|옥션|소개딜|티몬|자체쿠폰|삼카|\xa0","",article_title)
            product_title = product_title.replace(",", "")
            product_title = product_title.replace(".", "")
            product_title = product_title.replace('\n',"")
            product_title_list.append(product_title)
            
            ## 가격 추출 모듈
            re1 = re.compile('[\S\s]*([0-9]{6}원)[\S\s]*')
            re12 = re.compile('[\S\s]*([0-9]{5}원)[\S\s]*')
            re13 = re.compile('[\S\s]*([0-9]{4}원)[\S\s]*')
            re2 = re.compile('[\S\s]*([0-9]{6})[\S\s]*')
            re22 = re.compile('[\S\s]*([0-9]{5})[\S\s]*')
            re23 = re.compile('[\S\s]*([0-9]{4})[\S\s]*')
            priceRe = re1.match(product_title)
            priceRe12 = re12.match(product_title)
            priceRe13 = re13.match(product_title)
            priceRe2 = re2.match(product_title)
            priceRe22 = re22.match(product_title)
            priceRe23 = re23.match(product_title)
            if priceRe:
                price = priceRe.groups()[0]
                price = price.replace("원", "")
                print('1', price)
            elif priceRe12:
                price = priceRe12.groups()[0]
                price = price.replace("원", "")
                print('2', price)

            elif priceRe13:
                price = priceRe13.groups()[0]
                price = price.replace("원", "")
                print('2', price)
            elif priceRe2:
                price = priceRe2.groups()[0]
                price = price.replace("원", "")
                print('2', price)
            elif priceRe22:
                price = priceRe22.groups()[0]
                price = price.replace("원", "")
                print('2', price)
            elif priceRe23:
                price = priceRe23.groups()[0]
                price = price.replace("원", "")
                print('2', price)
            else:
                price = 999999
            price_list.append(price)
            
            # article_title.strip()
            # article_title = article_title.replace("\n", "") #Enter제거
            
            # 게시물링크
            # link = soup.select('tr.table_body > td.subject > div.relative > a.deco')[i]
            # link = link.attrs['href']
            # link = link.replace(" ","")
            link = "https://bbs.ruliweb.com/market/board/1020/read/"+str(article_num)+"?page="+str(page)
                
            
                
            # print(elements)
            matching_number_list.append(article_num)
            article_title_list.append(article_title)
            newurl_list.append(link)
            
            print(str(a+1)+"번째")
            print("id = "+str(article_num))
            print("제목 = "+article_title)
            print("링크 = "+link)
            
            a+=1
        
    
        if stop == 1:
            # worksheet_reset.insert_row(['%s월 %s일 %s시 %s분' % (now.month, now.day, now.hour, now.minute),'루리웹', reset_num, int(cell_num) + len(link_list)], 1)
            # print('%s월 %s일 %s시 %s분' % (now.month, now.day, now.hour, now.minute),'루리웹',reset_num, int(cell_num), len(link_list))
            break
    print(matching_number_list,article_title_list,newurl_list)
    
    
    # link_list -> 핫딜 상품을 실제로 구매하는 사이트 ex. 쿠팡 사이트
    
        
    for j in range(len(newurl_list)): 
        # 기본경로 설정  
        base_url = newurl_list[j]
        driver.implicitly_wait(5)
        
        html = requests.get(base_url)
        soup = bs(html.text, "html.parser")
        link = soup.select('div.source_url > a')
        
        # 링크가 존재하지 않을 때
        if len(link)==0:
            link = "링크없음"
        else:
            link = link[0].text
        
        
        print("링크는",link)
        link_list.append(link)
        
    print("전체링크는",link_list)


    # print('링크의 갯수는: ',len(link_list), '개')
    # worksheet_reset.insert_row(['%s월 %s일 %s시 %s분' % (now.month, now.day, now.hour, now.minute),'루리웹', reset_num, int(cell_num) + len(link_list)], 1)
    # print('%s월 %s일 %s시 %s분' % (now.month, now.day, now.hour, now.minute),'루리웹',reset_num, int(cell_num), len(link_list))
    
    print('==== reset_num inserted 1 ===== ')

                    
    insert_cols = ['A1', 'B1', 'C1', 'D1']
    insert_vals = ['%s월 %s일 %s시 %s분' % (now.month, now.day, now.hour, now.minute),'루리웹', reset_num, int(cell_num) + len(link_list)]

    load_wb.active = worksheet_reset 
    worksheet_reset.insert_rows(1,1)
    
    for (idx, val) in zip(insert_cols, insert_vals):
            print(idx, val)
            worksheet_reset[idx].value = val
        
    load_wb.save("./루리웹_데이터.xlsx")

    print('링크 리스트: ', link_list)
    print('[루리웹]링크의 갯수는: ',len(link_list), '개')

    # with open( 'temp_link.csv', 'w', newline='' ) as csvfile:
    #     spamwriter = csv.writer(csvfile)
    #     for link_num, newurl, article_title, product_title, link, price in zip(final_matching_list,newurl_list,article_title_list,product_title_list,link_list,price_list):
    #         spamwriter.writerow(['%s월 %s일 %s시 %s분' % (now.month, now.day, now.hour, now.minute), '뽐뿌', link_num, newurl,article_title, product_title, link, price])
    
    
    load_wb.active = worksheet 
    for link_num, newurl, article_title, product_title, link, price in zip(final_matching_list,newurl_list,article_title_list,product_title_list,link_list,price_list):
        worksheet.append(['%s월 %s일 %s시 %s분' % (now.month, now.day, now.hour, now.minute),'루리웹', link_num, newurl,article_title, product_title, link, price])
        print('엑셀 시트 작성중!!!')
        
    load_wb.save("./루리웹_데이터.xlsx")
    return len(link_list)

def shopmall(stop_num):
    global driver
    global cell_num
    global worksheet
    global worksheet_reset
    global load_wb
    # for i in range(300):
    #     cell_data = worksheet.acell( 'I' + str(i+1) ).value
    #     print(cell_data)
    #     if cell_data != None:
    #         stop_num = i
    #         print(stop_num)
    #         break
    print('링크내 제목 탐색 시작')
    print(cell_num,'부터 ',stop_num,'개를 탐색')
    if stop_num == 0:   # 할게 없으면 패스
        pass
    else:
        load_wb.active = worksheet  
        cell_link_list = worksheet['G' + str(cell_num) + ':G' + str(int(cell_num)+stop_num-1)]
        article_title_list=[]
        
        for i in cell_link_list:
            mall_link = i[0].value
            print(mall_link)
            try:
                driver.get(mall_link)
                # requests.get(mall_link)
            except:
                mall_title = '링크 접속불가'
                article_title_list.append(mall_title)
                continue
            sleep(10)
            # driver.implicitly_wait(10)
            try:
                if mall_link.find('auction') > 0:
                    try:
                        mall_title = driver.find_element_by_xpath('/html/body/div[9]/div[2]/div[2]/form/h1').text
                    except:
                        mall_title = driver.find_element_by_xpath('//*[@id="frmMain"]/h1]').text
        
                elif mall_link.find('lotteon') > 0:
                    try:
                        mall_title = driver.find_element_by_xpath(
                            '/html/body/div[1]/main/div/div[1]/div[2]/div[1]/div[2]/h1').text
                    except:
                        try:
                            mall_title = driver.find_element_by_xpath(
                                '/html/body/div[1]/main/div/div[1]/div[2]/div[1]/div[3]/h1').text
                        except:
                            mall_title = driver.find_element_by_xpath(
                                '//*[@id="stickyTopParent"]/div[2]/div[2]/div[3]/h1]').text
        
                elif 'wemakeprice' in mall_link:
                    try:
                        mall_title = driver.find_element_by_xpath(
                            '/html/body/div[2]/div[2]/div/div[1]/div[3]/div[1]/div[2]/div[1]/h3').text
                    except:
                        mall_title = driver.find_element_by_xpath('//*[@id="_infoDescription"]/div[1]/h3').text
        
                elif mall_link.find('gmarket') > 0:
                    try:
                        mall_title = driver.find_element_by_xpath(
                            '/html/body/div[4]/div[2]/div[2]/div[1]/div/div[1]/h1').text
                    except:
                        try:
                            mall_title = driver.find_element_by_xpath(
                                '/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div[2]/div[2]/div[1]/h3').text
                        except:
                            try:
                                mall_title = driver.find_element_by_xpath(
                                    '//*[@id="goodsDetailTab"]/div[2]/div/div[3]/h3').text
                            except:
                                mall_title = driver.find_element_by_xpath(
                                    '/html/body/div[1]/div[4]/div[2]/div[1]/div[2]/div/div[2]/div[1]/h3').text
        
                elif mall_link.find('gs') > 0:
                    try:
                        mall_title = driver.find_element_by_xpath('/html/body/div[4]/div[1]/section[1]/h1').text
                    except:
                        mall_title = driver.find_element_by_xpath('//*[@id="mainInfo"]/div[1]/section[1]/h1').text
        
                elif mall_link.find('tmon') > 0:
                    try:
                        mall_title = driver.find_element_by_xpath(
                            '/html/body/div[3]/div[2]/div/div/section[1]/section[1]/section[1]/div[3]/article[1]/div[2]/h2').text
                    except:
                        try:
                            mall_title = driver.find_element_by_xpath(
                                '/html/body/div[2]/section[1]/section[1]/div[4]/article[1]/div[2]/h2').text
                        except:
                            try:
                                mall_title = driver.find_element_by_xpath(
                                    '/html/body/div[2]/div/div/div/section[1]/section[1]/section[1]/div[3]/article[1]/div[2]/h2').text
                            except:
                                mall_title = driver.find_element_by_xpath(
                                    '/html/body/div[2]/div/div/div[1]/div[3]/div[1]/p[1]').text
                elif mall_link.find('11st') > 0:
                    try:
                        mall_title = driver.find_element_by_xpath(
                            '/html/body/div[2]/div[3]/div/div[1]/div[2]/div/div[1]/div[2]/div[2]/div[3]/h1').text
                    except:
                        try:
                            mall_title = driver.find_element_by_xpath(
                                '/html/body/div[2]/div[2]/div/div[1]/div[2]/div/div[1]/div[2]/div[2]/div[2]/h1').text
                        except:
                            mall_title = driver.find_element_by_xpath(
                                '//*[@id="layBodyWrap"]/div/div[1]/div[2]/div/div[1]/div[2]/div[2]/div[2]/h1').text
        
                elif mall_link.find('interpark') > 0:
                    mall_title = driver.find_element_by_xpath(
                        '/html/body/div[1]/div[5]/div/div[2]/div/div/div[2]/h2/span[2]').text
        
                elif mall_link.find('coupang') > 0:
                    mall_title = driver.find_element_by_xpath(
                        '/html/body/div[1]/section/div[1]/div/div[3]/div[3]/h2').text
        
                elif mall_link.find('naver') > 0:
                    try:
                        mall_title = driver.find_element_by_xpath(
                            '/html/body/div/div/div[2]/div[2]/div[2]/div[2]/fieldset/div[2]/div[1]/h3').text
                    except:
                        try:
                            mall_title = driver.find_element_by_xpath(
                                '/html/body/div/div/div[3]/div/div[1]/h2').text
                        except:
                            try:
                                mall_title = driver.find_element_by_xpath(
                                    '/html/body/div/div/div[3]/div[2]/div[2]/div/div[2]/div[2]/fieldset/div[1]/div[1]/h3').text
                            except:
                                mall_title = driver.find_element_by_xpath(
                                    '/html/body/div/div/div[2]/div/div[2]/div/div[2]/div[2]/fieldset/div[1]/div[2]/div/strong/span[2]').text
        
                elif mall_link.find('brand.naver') > 0:
                    mall_title = driver.find_element_by_xpath(
                        '/html/body/div/div/div[2]/div/div[2]/div/div[2]/div[2]/fieldset/div[1]/div[1]/h3').text
        
                elif mall_link.find('kakao') > 0:
                    mall_title = driver.find_element_by_xpath(
                        '/html/body/fu-app-root/fu-talkstore-wrapper/div/div/fu-pw-product-detail/div/div[1]/strong').text
        
                elif mall_link.find('yes24') > 0:
                    mall_title = driver.find_element_by_xpath(
                        '/html/body/div/div[4]/div[2]/div[1]/div/h2').text
        
                elif mall_link.find('nsmall') > 0:
                    mall_title = driver.find_element_by_xpath(
                        '/html/body/div/div[1]/div/div/section[1]/div[2]/div[1]/h3').text
        
                elif mall_link.find('ssg') > 0:
                    try:
                        mall_title = driver.find_element_by_xpath(
                            '/html/body/div[4]/div[6]/div[2]/div[1]/div[2]/div[2]/h2').text
                    except:
                        mall_title = driver.find_element_by_xpath(
                            '/html/body/div[4]/div[6]/div/div[2]/div[1]/div[2]/div[2]/h2').text
                else:
                    mall_title = "모르는 사이트"
            except:
                mall_title = "설정된 사이트, 설정안된 태그"
            article_title_list.append(mall_title)
            print('쇼핑몰에서 찾은 제품명: ',mall_title)
        print('전체 제품명 리스트는: ',article_title_list)
        cell_title_list = worksheet['I' + str(cell_num) + ':I' + str(int(cell_num) + stop_num-1)]
        for index, val in enumerate(article_title_list):
            cell_title_list[index][0].value = val
        # worksheet.update_cells(cell_title_list)
        load_wb.save("./루리웹_데이터.xlsx")
        
def navermall(stop_num):
    global driver
    global cell_num
    global worksheet
    global worksheet_reset
    global load_wb

    print('네이버에서 최저가 탐색 시작', stop_num, '개')
    if stop_num == 0:
        pass
    else:
        load_wb.active = worksheet
        cell_title_list = worksheet['I' + str(cell_num) + ':I' + str(int(cell_num)+stop_num-1)]
        # print(cell_title_list)
        '''index_cell의 숫자 확인해보기.'''
        for index_cell,mall_title in enumerate(cell_title_list):
            mall_title = mall_title[0].value
            url = "https://shopping.naver.com/"
            time.sleep(0.5)
            if mall_title == "설정된 사이트_설정안된 태그" : 
                mall_title = worksheet['F' + str(int(cell_num) + index_cell)].value
                
            elif mall_title == "모르는 사이트":
                mall_title = worksheet['F' + str(int(cell_num) + index_cell)].value
                
            elif mall_title == "링크 접속불가":
                mall_title = worksheet['F' + str(int(cell_num) + index_cell)].value

            if mall_title is None:
                continue
            driver.get(url)
                        
            element = driver.find_element_by_xpath('//*[@id="_verticalGnbModule"]/div/div[2]/div/div[2]/div/div[2]/form/fieldset/div/input')
            # 1. 검색창에 단어 입력하기
            print('네이버에서 탐색하는 제품명: ',mall_title)
            element.send_keys(mall_title)

            # driver.find_element_by_xpath('//*[@id="autocompleteWrapper"]/a[2]').click()
            driver.find_element_by_xpath('//*[@id="_verticalGnbModule"]/div/div[2]/div/div[2]/div/div[2]/form/fieldset/div/button[2]').click()

            time.sleep(2)

            # 3. 가격비교 창 들어가기
            try:
                diff = driver.find_element_by_xpath(
                    '//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/div[1]/ul/li[2]/a')
                diff.click()
            except:
                worksheet['J'+str(int(cell_num) + index_cell)].value = '네이버 가격비교 없음'
                # 특정 셀 쓰기
                continue
            time.sleep(1)
            try:
                naver_review_num = driver.find_element_by_xpath(
                    '/html/body/div/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[5]/a/em').text
            except:
                naver_review_num = '리뷰 없음'

            try:
                item = driver.find_element_by_xpath(
                    '//*[@id="__next"]/div/div[2]/div/div[3]/div[1]/ul/div/div[1]/li/div/div[2]/div[1]/a')

            # 가격 비교 상품 없을시
            except:
                worksheet['J'+str(int(cell_num) + index_cell)].value = '네이버 가격비교 없음'
                continue
            entry = item.get_attribute("data-nclick")
            entry = entry.split(",")
            catalog_id = entry[1].replace("i:", "")
            naver_link = 'https://search.shopping.naver.com/catalog/' + catalog_id
            driver.get(naver_link)
            time.sleep(1)
            # 배송료 포함버튼
            driver.find_element_by_xpath(
                '//*[@id="__next"]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/div/div/div/div[2]/span[1]/a').click()

            time.sleep(1)
            naver_pro_name = driver.find_element_by_xpath(
                '/html/body/div/div/div[2]/div[2]/div[1]/h2').text
            priceText = driver.find_element_by_xpath(
                '//*[@id="__next"]/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/table/tbody/tr[1]/td[2]/a/em').text
            priceText = priceText.replace(",", "")
            delivery = driver.find_element_by_xpath(
                '/html/body/div/div/div[2]/div[2]/div[2]/div[1]/div/div[2]/div[2]/table/tbody/tr[1]/td[3]').text

            all_list=[naver_link,catalog_id,naver_pro_name,priceText,delivery,naver_review_num]
            # cell_list_all = worksheet['J' + str(cell_num) + ':O' + str(int(cell_num)+stop_num-1)]
            
            insert_cols = ['J', 'K', 'L', 'M', 'N', 'O']

            for index, val in zip(insert_cols, all_list):
                worksheet[index + str(int(cell_num) + index_cell)].value = val

            load_wb.save("./루리웹_데이터.xlsx")
            #worksheet.update_cells(cell_list_all)


def checkFileExists( file ):
    for f in os.listdir('.'):
        if ( os.path.isfile(f) ) :
            if ( f == file ):
                return True
    return False

if ( checkFileExists( '루리웹_데이터.xlsx' ) == False ):
    write_wb = Workbook()
    write_wb.create_sheet('루리웹_기본정보')
    write_wb.create_sheet('루리웹_최신_게시글정보')
    write_wb['루리웹_기본정보'].append([
                    "시간", "핫딜몰", "게시글 id", "핫딜몰 링크", "핫딜몰 제품명", "보정 제품명", "쇼핑몰 주소", "핫딜몰 가격", 
                    "링크 접속불가", "네이버  주소", 
                    "네이버 번호", "네이버 제목", "네이버 가격", "네이버 배송료", "네이버 리뷰 개수"
                ])
    write_wb.save("./루리웹_데이터.xlsx")

    load_wb = load_workbook("./루리웹_데이터.xlsx")
    worksheet = load_wb['루리웹_기본정보']
    worksheet_reset = load_wb['루리웹_최신_게시글정보']

else:
    load_wb = load_workbook("./루리웹_데이터.xlsx")
    worksheet = load_wb['루리웹_기본정보']
    worksheet_reset = load_wb['루리웹_최신_게시글정보']

bigLoop = True

while bigLoop:
    try:
        row_cnt = int(len(list(worksheet_reset.rows)))
        print("row number is : ", row_cnt , " : ", worksheet_reset['A1'].value)

        if ( row_cnt == 0 ):
            searh_num = 0
            cell_num = 2
            print('탐색번호 init:',searh_num, cell_num)
        else:
            load_wb.active = worksheet_reset
            searh_num = worksheet_reset['C1'].value
            print('이전의 탐색번호:',searh_num)

            cell_num = worksheet_reset['D1'].value
            print('이전의 셀 번호', cell_num)
        
        now = datetime.now()
        option = Options()
        option.add_experimental_option("prefs", {"profile.default_content_setting_values.notifications": 1 })
        #option.add_experimental_option("excludeSwitches", ["enable-logging"])
        # option.add_argument('headless')

        if getattr(sys, 'frozen', False):
            chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
            driver = webdriver.Chrome(chromedriver_path)
        else:
            driver = webdriver.Chrome(options=option,executable_path='./chromedriver')

        #driver = webdriver.Chrome(options=option,executable_path='./chromedriver')

        new_search_num = hotdeal()

        if ( new_search_num == 0 ):
            print('최신 데이터임 30분휴식')
            time.sleep(3000)
            continue
        
        shopmall(new_search_num)

        navermall(new_search_num)
        
        # bigLoop = False
    except Exception as e:
        print('오류 발생, 10분 슬립')
        traceback.print_exc()

        driver.close()

        time.sleep(100)